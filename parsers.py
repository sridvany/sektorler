"""TCMB Sektör Bilançoları XLSX parser'ları - 6 tip."""
import io
import pandas as pd
from openpyxl import load_workbook


def _load(content: bytes):
    """XLSX bytes → (wb, ilk sheet). KARTIL/RISK sheet adı yıla göre değişiyor,
    o yüzden sheet adına değil ilk sheet'e güveniriz."""
    wb = load_workbook(io.BytesIO(content), data_only=True)
    ws = wb[wb.sheetnames[0]]
    return wb, ws


def _find_years(ws, search_rows=range(1, 12)):
    """İlk birkaç satırda 2000-2100 arası tam sayıları yıl olarak yakala."""
    found = []
    for r in search_rows:
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, int) and 2000 < v < 2100:
                if v not in found:
                    found.append(v)
    return sorted(found)[:2] if len(found) >= 2 else found


# ---------- BIL / GEL: hiyerarşik liste ----------
def parse_bil_gel(content: bytes):
    """B sütunu = etiket, C/D = yıl değerleri. AKTİF ve PASİF aynı sheet'te."""
    wb, ws = _load(content)
    meta = {
        "sector_name": ws.cell(row=2, column=1).value or "",
        "title": ws.cell(row=3, column=1).value or "",
    }
    years = _find_years(ws) or ["Yıl 1", "Yıl 2"]
    y1, y2 = str(years[0]), str(years[1]) if len(years) > 1 else "Yıl 2"

    rows = []
    for r in range(4, ws.max_row + 1):
        label = ws.cell(row=r, column=2).value
        v1 = ws.cell(row=r, column=3).value
        v2 = ws.cell(row=r, column=4).value
        if label is None and v1 is None and v2 is None:
            continue
        # Yıl başlığı satırını atla (AKTİF/PASİF iki kez geçtiği için)
        if isinstance(v1, int) and 2000 < v1 < 2100:
            continue
        rows.append({"Kalem": label, y1: v1, y2: v2})

    return pd.DataFrame(rows), meta


# ---------- YAPISAL ----------
def parse_yapisal(content: bytes):
    """Her yıl için tutar + (%) yan yana. 3 bölüm: Aktif/Pasif/Öz Kaynak yapısı."""
    wb, ws = _load(content)
    meta = {
        "sector_name": ws.cell(row=2, column=1).value or "",
        "title": ws.cell(row=3, column=1).value or "",
    }
    # r5: [None, None, '2023', '(%)', '2024', '(%)']
    years_raw = [ws.cell(row=5, column=3).value, ws.cell(row=5, column=5).value]
    years = []
    for y in years_raw:
        try:
            years.append(int(y))
        except (TypeError, ValueError):
            years.append(y if y else "?")
    y1, y2 = str(years[0]), str(years[1])

    rows = []
    current_section = ""
    for r in range(6, ws.max_row + 1):
        label = ws.cell(row=r, column=2).value
        v1 = ws.cell(row=r, column=3).value
        p1 = ws.cell(row=r, column=4).value
        v2 = ws.cell(row=r, column=5).value
        p2 = ws.cell(row=r, column=6).value
        if label is None:
            continue
        # Bölüm başlığı: değer yok, etiket var
        if v1 is None and v2 is None and p1 is None and p2 is None:
            current_section = str(label).strip()
            continue
        rows.append({
            "Bölüm": current_section,
            "Kalem": label,
            f"{y1} (Bin TL)": v1,
            f"{y1} (%)": p1,
            f"{y2} (Bin TL)": v2,
            f"{y2} (%)": p2,
        })
    return pd.DataFrame(rows), meta


# ---------- KARTIL (Oranlar) ----------
def parse_kartil(content: bytes):
    """Her yıl için Firma Sayısı, Q (ağırlıklı), Q1, Q2, Q3."""
    wb, ws = _load(content)
    meta = {
        "sector_name": ws.cell(row=2, column=1).value or "",
        "title": ws.cell(row=3, column=1).value or "",
    }
    years = _find_years(ws, range(1, 9)) or ["Yıl 1", "Yıl 2"]
    if len(years) < 2:
        years = years + ["Yıl 2"]
    y1, y2 = str(years[0]), str(years[1])

    rows = []
    current_category = ""
    for r in range(9, ws.max_row + 1):
        cat = ws.cell(row=r, column=1).value
        label = ws.cell(row=r, column=2).value
        # Kategori başlığı (örn. "A-" + " LİKİDİTE ORANLARI")
        if cat and label and not str(cat).strip().startswith("*"):
            current_category = str(label).strip()
            continue
        if label is None or not str(label).strip():
            continue
        if str(label).strip().startswith("*"):
            continue
        rows.append({
            "Kategori": current_category,
            "Oran": str(label).strip(),
            f"{y1} Firma": ws.cell(row=r, column=3).value,
            f"{y1} Q (Ağr.)": ws.cell(row=r, column=4).value,
            f"{y1} Q1": ws.cell(row=r, column=5).value,
            f"{y1} Q2 (Med.)": ws.cell(row=r, column=6).value,
            f"{y1} Q3": ws.cell(row=r, column=7).value,
            f"{y2} Firma": ws.cell(row=r, column=8).value,
            f"{y2} Q (Ağr.)": ws.cell(row=r, column=9).value,
            f"{y2} Q1": ws.cell(row=r, column=10).value,
            f"{y2} Q2 (Med.)": ws.cell(row=r, column=11).value,
            f"{y2} Q3": ws.cell(row=r, column=12).value,
        })
    return pd.DataFrame(rows), meta


# ---------- RISK ----------
def parse_risk(content: bytes):
    """Matris: Yıl × Vade (Kısa/Uzun/Toplam) × Para (TL/YP).
    Sektör adı r4'te, başlık r6 veya r8'de."""
    wb, ws = _load(content)
    sector_name = ws.cell(row=4, column=2).value or ""
    title = ws.cell(row=6, column=2).value or ws.cell(row=8, column=2).value or ""
    meta = {"sector_name": sector_name, "title": title}

    years = _find_years(ws, range(1, 11)) or ["Yıl 1", "Yıl 2"]
    if len(years) < 2:
        years = years + ["Yıl 2"]
    y1, y2 = str(years[0]), str(years[1])

    # Data r11'den itibaren. Sütunlar: c4-c6 = y1 KV/UV/Toplam, c7-c9 = y2 KV/UV/Toplam
    rows = []
    for r in range(11, ws.max_row + 1):
        label = ws.cell(row=r, column=2).value
        if not label or not str(label).strip():
            continue
        rows.append({
            "Kalem": str(label).strip(),
            f"{y1} Kısa Vadeli": ws.cell(row=r, column=4).value,
            f"{y1} Uzun Vadeli": ws.cell(row=r, column=5).value,
            f"{y1} Toplam": ws.cell(row=r, column=6).value,
            f"{y2} Kısa Vadeli": ws.cell(row=r, column=7).value,
            f"{y2} Uzun Vadeli": ws.cell(row=r, column=8).value,
            f"{y2} Toplam": ws.cell(row=r, column=9).value,
        })
    return pd.DataFrame(rows), meta


# ---------- SEKKIM (Sektör Kimliği) ----------
def parse_sekkim(content: bytes):
    """7 alt blok içerir. dict[str, DataFrame] döner."""
    wb, ws = _load(content)

    def cv(r, c):
        return ws.cell(row=r, column=c).value

    blocks = {}

    # 1-3: Sektör, Yıl-Dönem, Firma Sayısı (r6, r8, r10)
    blocks["info"] = pd.DataFrame([
        {"Alan": "Sektör", "Değer": cv(6, 5)},
        {"Alan": "Yıl ve Dönem", "Değer": cv(8, 5)},
        {"Alan": "İncelenen Firma Sayısı", "Değer": cv(10, 5)},
    ])

    # 4: Hukuki yapı (r13-r21 arasında tara)
    legal_rows = []
    for r in range(13, 22):
        lbl, val = cv(r, 2), cv(r, 5)
        if lbl and val is not None:
            legal_rows.append({"Hukuki Yapı": str(lbl).strip(), "Firma Sayısı": val})
    blocks["legal"] = pd.DataFrame(legal_rows)

    # 5: Ölçek dağılımı (r25-r30)
    scale_rows = []
    for r in range(25, 31):
        lbl = cv(r, 2)
        if not lbl:
            continue
        scale_rows.append({
            "Ölçek": str(lbl).strip(),
            "Firma Sayısı": cv(r, 4),
            "Çalışan Sayısı": cv(r, 5),
            "Çalışan %": cv(r, 6),
            "Net Satışlar (Bin TL)": cv(r, 7),
            "Net Satışlar %": cv(r, 8),
            "Aktif (Bin TL)": cv(r, 9),
            "Aktif %": cv(r, 10),
            "Öz Kaynaklar (Bin TL)": cv(r, 11),
            "Öz Kaynaklar %": cv(r, 12),
        })
    blocks["scale"] = pd.DataFrame(scale_rows)

    # 6: Sektör Riski (yıllar r33'te c7 ve c10, data r35-r49)
    years = []
    for c in range(1, ws.max_column + 1):
        v = cv(33, c)
        if isinstance(v, int) and 2000 < v < 2100:
            years.append(v)
    if len(years) < 2:
        years = (years + ["Yıl 1", "Yıl 2"])[:2]
    y1, y2 = str(years[0]), str(years[1])

    risk_rows = []
    for r in range(35, 50):
        lbl = cv(r, 2)
        if not lbl:
            continue
        risk_rows.append({
            "Kalem": str(lbl).strip(),
            f"{y1} Kısa Vadeli": cv(r, 7),
            f"{y1} Uzun Vadeli": cv(r, 8),
            f"{y1} Toplam": cv(r, 9),
            f"{y2} Kısa Vadeli": cv(r, 10),
            f"{y2} Uzun Vadeli": cv(r, 11),
            f"{y2} Toplam": cv(r, 12),
        })
    blocks["risk"] = pd.DataFrame(risk_rows)

    # 7: Bilanço dönemi sonuçları (r52-r55)
    period_rows = []
    for r in range(52, 56):
        lbl, val = cv(r, 2), cv(r, 4)
        if lbl and val is not None:
            period_rows.append({"Durum": str(lbl).strip(), "Firma Sayısı": val})
    blocks["period"] = pd.DataFrame(period_rows)

    meta = {"sector_name": cv(6, 5) or "", "title": "SEKTÖR KİMLİĞİ"}
    return blocks, meta


# ---------- Dispatcher ----------
PARSERS = {
    "BIL": parse_bil_gel,
    "GEL": parse_bil_gel,
    "YAPISAL": parse_yapisal,
    "KARTIL": parse_kartil,
    "RISK": parse_risk,
    "SEKKIM": parse_sekkim,
}


def parse(file_type: str, content: bytes):
    if file_type not in PARSERS:
        raise ValueError(f"Bilinmeyen dosya tipi: {file_type}")
    return PARSERS[file_type](content)
